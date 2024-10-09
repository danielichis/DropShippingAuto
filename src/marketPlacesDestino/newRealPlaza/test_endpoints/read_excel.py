from openpyxl import load_workbook
workbook = load_workbook(filename='file.xlsx')
#get first sheet index
first_sheet = workbook.sheetnames[0]
last_row = workbook[first_sheet].max_row
last_column = workbook[first_sheet].max_column
print("last row",last_row)
print("last column",last_column)
# # Ahora puedes trabajar con el workbook como desees
# print(workbook.sheetnames)
# # read cell A1 IN FIRST SHEET
# print(workbook.active['A1'].value)
# # READ CELL 2 ROW 18 COLUMN IN FIRST SHEET
# print(workbook.active.cell(row=2, column=25).value)
# print(workbook.active.cell(row=2, column=25).comment.text)
